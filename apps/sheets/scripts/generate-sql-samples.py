#!/usr/bin/env python3
"""Generate the two ZenOffice SQL workbook samples and their shared test manifest."""

from __future__ import annotations

import json
import shutil
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "fixtures" / "sql"
OUTPUT.mkdir(parents=True, exist_ok=True)
WEB_OUTPUT = ROOT.parents[1] / "docs" / "samples"
WEB_OUTPUT.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10, color="1F2937")
THIN = Side(style="thin", color="D8E2EA")
BODY_BORDER = Border(bottom=THIN)


def rows(headers, records):
    return [headers] + [[record.get(header) for header in headers] for record in records]


def style_sheet(ws, table_name):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(vertical="center")
    for column_cells in ws.columns:
        values = [str(cell.value or "") for cell in column_cells[:80]]
        width = min(36, max(10, max((len(value) for value in values), default=8) + 2))
        ws.column_dimensions[column_cells[0].column_letter].width = width
    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def build_workbook(path, sheets, schema):
    wb = Workbook()
    wb.remove(wb.active)
    for index, (sheet_name, matrix) in enumerate(sheets.items(), start=1):
        ws = wb.create_sheet(sheet_name)
        for row in matrix:
            ws.append(row)
        style_sheet(ws, f"GenOfficeTable{index}")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
    meta = wb.create_sheet("_GenOfficeSchema")
    meta["A1"] = json.dumps({"version": 1, "tables": schema}, ensure_ascii=False, separators=(",", ":"))
    meta.sheet_state = "veryHidden"
    wb.save(path)


def retail_sample():
    cities = ["上海", "北京", "深圳", "杭州", "成都", "武汉"]
    customers = []
    for i in range(1, 13):
        customers.append({
            "客户ID": f"C{i:03d}",
            "客户名称": ["星河科技", "远山贸易", "海棠书店", "青禾教育", "极光设计", "启明医疗", "云帆物流", "拾光咖啡", "澄海制造", "新芽生物", "山川文创", "融汇咨询"][i - 1],
            "城市": cities[(i - 1) % len(cities)],
            "客户等级": ["A", "B", "C"][i % 3],
            "注册日期": date(2023, 1, 12) + timedelta(days=i * 41),
            "联系电话": None if i in (4, 10) else f"1380000{i:04d}",
        })
    products = []
    product_names = [
        ("P001", "人体工学键盘", "办公外设", 399.0),
        ("P002", "静音鼠标", "办公外设", 159.0),
        ("P003", "4K 显示器", "显示设备", 2699.0),
        ("P004", "移动固态硬盘", "存储设备", 699.0),
        ("P005", "降噪耳机", "音频设备", 1299.0),
        ("P006", "视频会议摄像头", "会议设备", 899.0),
        ("P007", "USB-C 扩展坞", "办公外设", 499.0),
        ("P008", "激光打印机", "打印设备", 1899.0),
        ("P009", "可调节显示器支架", "办公家具", 329.0),
        ("P010", "智能白板", "会议设备", 5899.0),
    ]
    for idx, (pid, name, category, price) in enumerate(product_names):
        products.append({"商品ID": pid, "商品名称": name, "类别": category, "标准单价": price, "是否在售": idx != 7, "上市日期": date(2022, 6, 1) + timedelta(days=idx * 67)})

    orders = []
    order_lines = []
    for i in range(1, 31):
        oid = f"O2025{i:04d}"
        customer = customers[(i * 5) % len(customers)]["客户ID"]
        order_date = date(2025, 1, 3) + timedelta(days=i * 11)
        orders.append({
            "订单ID": oid,
            "客户ID": customer,
            "订单日期": order_date,
            "销售人员": ["陈晨", "李妍", "王睿", "赵宁"][i % 4],
            "渠道": ["直营网店", "经销商", "企业采购"][i % 3],
            "订单状态": "已取消" if i in (9, 23) else ("待发货" if i > 27 else "已完成"),
            "收货省份": ["上海", "北京", "广东", "浙江", "四川", "湖北"][i % 6],
        })
        line_count = 3 if i % 4 == 0 else 2
        for line_no in range(1, line_count + 1):
            product = products[(i * 3 + line_no * 2) % len(products)]
            quantity = 1 + (i + line_no) % 5
            discount = [1.0, 0.95, 0.9, 0.85][(i + line_no) % 4]
            unit_price = product["标准单价"]
            order_lines.append({
                "订单ID": oid,
                "行号": line_no,
                "商品ID": product["商品ID"],
                "数量": quantity,
                "成交单价": unit_price,
                "折扣率": discount,
                "行金额": round(quantity * unit_price * discount, 2),
                "备注": "加急" if i % 10 == 0 and line_no == 1 else None,
            })

    matrices = {
        "客户": rows(["客户ID", "客户名称", "城市", "客户等级", "注册日期", "联系电话"], customers),
        "订单": rows(["订单ID", "客户ID", "订单日期", "销售人员", "渠道", "订单状态", "收货省份"], orders),
        "订单明细": rows(["订单ID", "行号", "商品ID", "数量", "成交单价", "折扣率", "行金额", "备注"], order_lines),
        "商品": rows(["商品ID", "商品名称", "类别", "标准单价", "是否在售", "上市日期"], products),
    }
    schema = [
        {"sheetName": "客户", "tableName": "客户", "primaryKey": ["客户ID"], "nullable": ["联系电话"], "types": {"客户ID": "TEXT", "客户名称": "TEXT", "城市": "TEXT", "客户等级": "TEXT", "注册日期": "DATE", "联系电话": "TEXT"}, "indexes": [{"name": "idx_客户_城市等级", "columns": ["城市", "客户等级"], "unique": False}]},
        {"sheetName": "订单", "tableName": "订单", "primaryKey": ["订单ID"], "nullable": [], "types": {"订单ID": "TEXT", "客户ID": "TEXT", "订单日期": "DATE", "销售人员": "TEXT", "渠道": "TEXT", "订单状态": "TEXT", "收货省份": "TEXT"}, "indexes": [{"name": "idx_订单_客户日期", "columns": ["客户ID", "订单日期"], "unique": False}]},
        {"sheetName": "订单明细", "tableName": "订单明细", "primaryKey": ["订单ID", "行号"], "nullable": ["备注"], "types": {"订单ID": "TEXT", "行号": "INTEGER", "商品ID": "TEXT", "数量": "INTEGER", "成交单价": "NUMBER", "折扣率": "NUMBER", "行金额": "NUMBER", "备注": "TEXT"}, "indexes": [{"name": "idx_明细_商品", "columns": ["商品ID"], "unique": False}]},
        {"sheetName": "商品", "tableName": "商品", "primaryKey": ["商品ID"], "nullable": [], "types": {"商品ID": "TEXT", "商品名称": "TEXT", "类别": "TEXT", "标准单价": "NUMBER", "是否在售": "BOOLEAN", "上市日期": "DATE"}, "indexes": [{"name": "idx_商品_类别在售", "columns": ["类别", "是否在售"], "unique": False}]},
    ]
    queries = [
        {"name": "城市销售额", "sql": "SELECT c.[城市], ROUND(SUM(d.[行金额]),2) AS [销售额] FROM [客户] c JOIN [订单] o ON o.[客户ID]=c.[客户ID] JOIN [订单明细] d ON d.[订单ID]=o.[订单ID] WHERE o.[订单状态]<>'已取消' GROUP BY c.[城市] ORDER BY [销售额] DESC", "min_rows": 6},
        {"name": "月度销售趋势", "sql": "SELECT YEAR(o.[订单日期]) AS [年份], MONTH(o.[订单日期]) AS [月份], ROUND(SUM(d.[行金额]),2) AS [销售额] FROM [订单] o JOIN [订单明细] d ON d.[订单ID]=o.[订单ID] WHERE o.[订单状态]<>'已取消' GROUP BY YEAR(o.[订单日期]), MONTH(o.[订单日期]) ORDER BY [年份],[月份]", "min_rows": 10},
        {"name": "商品销量", "sql": "SELECT p.[类别], SUM(d.[数量]) AS [销量], ROUND(SUM(d.[行金额]),2) AS [销售额] FROM [商品] p JOIN [订单明细] d ON d.[商品ID]=p.[商品ID] GROUP BY p.[类别] ORDER BY [销售额] DESC", "min_rows": 5},
    ]
    return matrices, schema, queries


def research_sample():
    researchers = []
    names = ["徐敏", "李哲", "王璐", "赵天宇", "陈思远", "周宁", "孙悦", "吴昊", "郑可", "何清", "蒋凡", "马骁"]
    for i, name in enumerate(names, start=1):
        researchers.append({
            "人员ID": f"R{i:03d}", "姓名": name,
            "职称": ["教授", "副教授", "研究员", "博士后"][i % 4],
            "机构": ["复旦大学", "上海交通大学", "浙江大学", "中国科学院"][i % 4],
            "邮箱": f"researcher{i}@example.edu.cn",
            "ORCID": None if i in (5, 9, 12) else f"0000-0002-{1000+i:04d}-{2000+i:04d}",
        })
    projects = []
    project_names = ["多模态科研智能体", "可信医学大模型", "超材料逆向设计", "低碳城市数字孪生", "量子传感算法", "柔性机器人控制", "生物信息知识图谱", "高性能存算一体芯片"]
    for i, name in enumerate(project_names, start=1):
        projects.append({
            "项目ID": f"PRJ{i:03d}", "项目名称": name,
            "项目类型": ["国家自然科学基金", "重点研发计划", "企业联合项目"][i % 3],
            "批准日期": date(2022, 1, 1) + timedelta(days=i * 113),
            "预算万元": float(40 + i * 17),
            "状态": ["在研", "在研", "结题"][i % 3],
            "负责人ID": researchers[(i * 2) % len(researchers)]["人员ID"],
        })
    members = []
    for i, project in enumerate(projects):
        for offset in range(4):
            researcher = researchers[(i * 3 + offset) % len(researchers)]
            members.append({
                "项目ID": project["项目ID"], "人员ID": researcher["人员ID"],
                "成员角色": ["负责人", "骨干", "参与人", "研究助理"][offset],
                "加入日期": project["批准日期"] + timedelta(days=offset * 19),
                "工时占比": [0.35, 0.25, 0.20, 0.15][offset],
            })
    papers = []
    for i in range(1, 21):
        project_id = None if i in (6, 17) else projects[(i * 3) % len(projects)]["项目ID"]
        papers.append({
            "论文ID": f"PUB{i:03d}",
            "项目ID": project_id,
            "论文标题": f"面向复杂场景的智能方法研究（第 {i} 篇）",
            "发表日期": date(2023, 2, 1) + timedelta(days=i * 47),
            "期刊会议": ["IEEE TPAMI", "Nature Communications", "Science Advances", "NeurIPS", "Cell Systems"][i % 5],
            "DOI": f"10.1234/genoffice.2025.{i:04d}",
            "通讯作者ID": researchers[(i * 5) % len(researchers)]["人员ID"],
            "引用次数": (i * 7) % 43,
        })
    matrices = {
        "项目": rows(["项目ID", "项目名称", "项目类型", "批准日期", "预算万元", "状态", "负责人ID"], projects),
        "研究人员": rows(["人员ID", "姓名", "职称", "机构", "邮箱", "ORCID"], researchers),
        "论文": rows(["论文ID", "项目ID", "论文标题", "发表日期", "期刊会议", "DOI", "通讯作者ID", "引用次数"], papers),
        "项目成员": rows(["项目ID", "人员ID", "成员角色", "加入日期", "工时占比"], members),
    }
    schema = [
        {"sheetName": "项目", "tableName": "项目", "primaryKey": ["项目ID"], "nullable": [], "types": {"项目ID": "TEXT", "项目名称": "TEXT", "项目类型": "TEXT", "批准日期": "DATE", "预算万元": "NUMBER", "状态": "TEXT", "负责人ID": "TEXT"}, "indexes": [{"name": "idx_项目_类型状态", "columns": ["项目类型", "状态"], "unique": False}]},
        {"sheetName": "研究人员", "tableName": "研究人员", "primaryKey": ["人员ID"], "nullable": ["ORCID"], "types": {"人员ID": "TEXT", "姓名": "TEXT", "职称": "TEXT", "机构": "TEXT", "邮箱": "TEXT", "ORCID": "TEXT"}, "indexes": [{"name": "idx_人员_机构职称", "columns": ["机构", "职称"], "unique": False}]},
        {"sheetName": "论文", "tableName": "论文", "primaryKey": ["论文ID"], "nullable": ["项目ID"], "types": {"论文ID": "TEXT", "项目ID": "TEXT", "论文标题": "TEXT", "发表日期": "DATE", "期刊会议": "TEXT", "DOI": "TEXT", "通讯作者ID": "TEXT", "引用次数": "INTEGER"}, "indexes": [{"name": "idx_论文_项目日期", "columns": ["项目ID", "发表日期"], "unique": False}, {"name": "uq_论文_DOI", "columns": ["DOI"], "unique": True}]},
        {"sheetName": "项目成员", "tableName": "项目成员", "primaryKey": ["项目ID", "人员ID"], "nullable": [], "types": {"项目ID": "TEXT", "人员ID": "TEXT", "成员角色": "TEXT", "加入日期": "DATE", "工时占比": "NUMBER"}, "indexes": [{"name": "idx_成员_人员", "columns": ["人员ID"], "unique": False}]},
    ]
    queries = [
        {"name": "项目团队与预算", "sql": "SELECT p.[项目名称], p.[预算万元], COUNT(m.[人员ID]) AS [成员数] FROM [项目] p JOIN [项目成员] m ON m.[项目ID]=p.[项目ID] GROUP BY p.[项目ID],p.[项目名称],p.[预算万元] ORDER BY p.[预算万元] DESC", "min_rows": 8},
        {"name": "人员跨项目参与", "sql": "SELECT r.[姓名],r.[机构],COUNT(m.[项目ID]) AS [参与项目数] FROM [研究人员] r JOIN [项目成员] m ON m.[人员ID]=r.[人员ID] GROUP BY r.[人员ID],r.[姓名],r.[机构] HAVING COUNT(m.[项目ID])>=2 ORDER BY [参与项目数] DESC", "min_rows": 4},
        {"name": "缺失项目归属论文", "sql": "SELECT [论文ID],[论文标题],[DOI] FROM [论文] WHERE [项目ID] IS NULL", "rows": 2},
        {"name": "年度论文引用", "sql": "SELECT YEAR([发表日期]) AS [年份],COUNT(*) AS [论文数],SUM([引用次数]) AS [总引用] FROM [论文] GROUP BY YEAR([发表日期]) ORDER BY [年份]", "min_rows": 2},
    ]
    return matrices, schema, queries


def json_value(value):
    return value.isoformat() if isinstance(value, date) else value


def main():
    samples = {}
    for key, filename, builder in [
        ("retail", "零售订单数据库.xlsx", retail_sample),
        ("research", "科研项目数据库.xlsx", research_sample),
    ]:
        matrices, schema, queries = builder()
        build_workbook(OUTPUT / filename, matrices, schema)
        shutil.copy2(OUTPUT / filename, WEB_OUTPUT / filename)
        samples[key] = {
            "file": filename,
            "schema": schema,
            "sheets": {name: [[json_value(value) for value in row] for row in matrix] for name, matrix in matrices.items()},
            "queries": queries,
        }
    (OUTPUT / "sql-samples.manifest.json").write_text(json.dumps({"version": 1, "samples": samples}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"status": "ok", "output": str(OUTPUT), "web_output": str(WEB_OUTPUT), "files": [sample["file"] for sample in samples.values()]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
